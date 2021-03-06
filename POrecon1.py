

import openpyxl   #Import OpenPyxl module
wb = openpyxl.load_workbook('VCE.xlsx')               #Open the excel file to be operated on
sheet = wb.get_sheet_by_name('Report')                  #Open the sub sheet "Report"
listpo = []             #Create a new list to store PO's
          
i = 3                   #Scrap Sheet data from the third row
while(1==1):            #Keep going unless stopped by a break function
    i = i + 1           #Keep increasing the Rows by 1
    
    cellvalueqad = sheet['X'+str(i)].value    #Store the QAD PO value in a variable 
    cellvaluevss = sheet['A'+str(i)].value     #Store the VSS PO value in a variable
    if(sheet['R'+str(i)].value == None):      #If no value found in the R column of sheet break and stop the while loop
        break
    listpo.append(cellvaluevss)
    if(cellvaluevss == None):
        listpo.append(cellvalueqad)



             
nb = openpyxl.load_workbook('Reconc.xlsx')       #open recon file to save data




def removedupli(seq):               #Remove duplicate POs from listpo 
    seen = set()                
    seen_add = seen.add
    return [ x for x in seq if not (x in seen or seen_add(x))]

povssdata = []          #Store VSS data in a list to write in excel
poqaddata = []          #Store QAD data in a list to write in excel
ponum = ''          
pops = ''               
poqty = 0

if listpo:                  #if any items in listpo
    listpo1 = removedupli(listpo)   #Remove duplicates from list PO and store in new variable
    for value in listpo1:
        i = 3                   #Start from 3rd row
        sumvssqty = 0
        sumqadqty = 0
        while(1==1):                #Keep continuing unless there is a break
            
            i = i + 1
            cellvalueqad = str(sheet['X'+str(i)].value)         #Store PO value in QAD in a variable
            cellvaluevss = str(sheet['A'+str(i)].value)          #Store PO value in QAD in a variable
            
            if(sheet['R'+str(i)].value == None):        #Stop when R column of excel sheet is none
                break
            if cellvaluevss==value:                     #For every PO in listpo1 get the data from corresponding columns in VSS and store in povssdata[] list

                if(sheet['E'+str(i)].value!=None):
                    pops = str(sheet['E'+str(i)].value.encode('utf-8'))
                else:
                    pops = str(sheet['E'+str(i)].value)
                poqty = sheet['F'+str(i)].value

                if(poqty!=None):
                    sumvssqty = sumvssqty + poqty
    
                povssdata.append([value,[pops,poqty,sumvssqty]])

                if(sheet['AB'+str(i)].value!=None):
                    pops = str(sheet['AB'+str(i)].value.encode('utf-8'))
                else:
                    pops = str(sheet['AB'+str(i)].value)
                poqty = sheet['AC'+str(i)].value
                
                if(poqty!=None):
                    sumqadqty = sumqadqty + poqty
                poqaddata.append([value,[pops,poqty,sumqadqty]])

            elif cellvalueqad==value:                   #For every PO in listpo1 get the data from corresponding columns in QAD and store in poqaddata[] list

                if(sheet['E'+str(i)].value!=None):
                    pops = str(sheet['E'+str(i)].value.encode('utf-8'))
                else:
                    pops = str(sheet['E'+str(i)].value)
                poqty = sheet['F'+str(i)].value

                if(poqty!=None):
                    sumvssqty = sumvssqty + poqty
                povssdata.append([value,[pops,poqty,sumvssqty]])

                if(sheet['AB'+str(i)].value!=None):
                    pops = str(sheet['AB'+str(i)].value.encode('utf-8'))
                else:
                    pops = str(sheet['AB'+str(i)].value)
                poqty = sheet['AC'+str(i)].value
                
                if(poqty!=None):
                    sumqadqty = sumqadqty + poqty
                poqaddata.append([value,[pops,poqty,sumqadqty]])

        


nb.create_sheet(index=0, title='VCE')                 #Create a new sheet in recon file with the name
sheetnew = nb.get_sheet_by_name('VCE')                #Select the newly created sheet


sheetnew['A1'] = 'Purchase Order'
sheetnew['A1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(povssdata)):
    sheetnew['A'+str(i+2)] = povssdata[i][0]                #Save VSS PO data in the sheet
    

sheetnew['B1'] = 'Qty missing in QAD'
sheetnew['B1'].font = openpyxl.styles.Font(size=12,bold=True)


for i in range(0,len(povssdata)):
    if(poqaddata[i][1][1] == None):
        sheetnew['B'+str(i+2)] = povssdata[i][1][1]                     
    else:
        sheetnew['B'+str(i+2)] = povssdata[i][1][1]-poqaddata[i][1][1]           #Save difference in qty data in the sheet

    
sheetnew['C1'] = 'Packaging Slip Number in VSS'
sheetnew['C1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(povssdata)):
    sheetnew['C'+str(i+2)] = povssdata[i][1][0]                      #Save VSS Packing Slip data in the sheet
    
sheetnew['D1'] = 'QTY In vSS'
sheetnew['D1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(povssdata)):
    sheetnew['D'+str(i+2)] = povssdata[i][1][1]              #Save VSS qty data in the sheet
    
sheetnew['E1'] = 'Packaging Slip Number in QAD'
sheetnew['E1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(poqaddata)):
    sheetnew['E'+str(i+2)] = poqaddata[i][1][0]                 #Save QAD Packing Slip data in the sheet

sheetnew['F1'] = 'QTY in QAD'
sheetnew['F1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(poqaddata)):
    sheetnew['F'+str(i+2)] = poqaddata[i][1][1]               #Save QAD qty data in the sheet

totalvss = []    
sheetnew['G1'] = 'Total in VSS'
sheetnew['G1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(poqaddata)):
    if (sheetnew['A'+str(i+3)].value!= sheetnew['A'+str(i+2)].value):
        totalvss.append(povssdata[i][1][2])

totalqad =[]

sheetnew['H1'] = 'Total in QAD'
sheetnew['H1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(poqaddata)):
    if (sheetnew['A'+str(i+3)].value!= sheetnew['A'+str(i+2)].value):
        totalqad.append(poqaddata[i][1][2])
        
totalmissing = []

sheetnew['I1'] = 'Total Missing in QAD'
sheetnew['I1'].font = openpyxl.styles.Font(size=12,bold=True)

for i in range(0,len(poqaddata)):
    if (sheetnew['A'+str(i+3)].value!= sheetnew['A'+str(i+2)].value):
        totalmissing.append(povssdata[i][1][2]-poqaddata[i][1][2])


mstart = 2
mlist = []
mlist.append(mstart)
for i in range(0,len(povssdata)):
    if (sheetnew['A'+str(i+3)].value!=sheetnew['A'+str(i+2)].value):
        sheetnew.merge_cells('A'+str(mstart)+':A'+str(i+2))
        sheetnew.merge_cells('G'+str(mstart)+':G'+str(i+2))
        sheetnew.merge_cells('H'+str(mstart)+':H'+str(i+2))
        sheetnew.merge_cells('I'+str(mstart)+':I'+str(i+2))
        mstart = i+3
        mlist.append(mstart)
   
for i in range(0,len(totalvss)):
    sheetnew['G'+str(mlist[i])].value = totalvss[i]
    sheetnew['H'+str(mlist[i])].value = totalqad[i]
    sheetnew['I'+str(mlist[i])].value = totalmissing[i]
    
nb.save('Reconc.xlsx')
print "finished"


